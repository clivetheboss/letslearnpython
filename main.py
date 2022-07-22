# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.


# def print_hi(name):
# Use a breakpoint in the code line below to debug your script.
# print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
# if __name__ == '__main__':
# print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
# variables

# price = 10
# print(price)

full_name = "John Smith"
age = 20
is_new = True
# print(age)

# name = input("What is your name? ")
# favorite_color = input("What is your favorite color")
# print("Hi" + name)
# print(name + " likes " + favorite_color )

# Calculating birth_year
# birth_year = input("Birth year: ")
# age = 2022 - int(birth_year)
# print(type(age))
# print(age)

# weight conversation example

# weight_lbs = input('Weight (lbs): ')
# weight_kg = int(weight_lbs) * 0.45
# print(weight_kg)

# learning strings----------------------

# course = "Python's course for Beginners"
# print(course)

# course = '''
# Hi Clive,

# Here is our first email to you

# Thank you

# '''
# print(course)

# index--------------------------------

# course = "Python's course for Beginners"
# print(course[0:3])
# -1 from last : all characters

# first = "Clive"
# last = "Madungwe"
# message = first + " [" + last + "] is a coder"
# msg = f"{first} [{last}] is a coder"
# print(message)
# print(msg)

# calculating length of strings---------------
# course = "Python for Beginners"
# print(len(course))
# print(course.upper()) # explore methods find, replace

# Arithmetic operations
# print(10/3) # // get integer
# print(10 ** 3) #exponent

# x = 10
# x = x + 3
# x += 3 # -=
# print(x)

# x = 10 + 3 * 2
# print(x)

# math function-------------

# x = 2.9
# print(round(x))

# x = -2.9
# print(abs(-2.9))

# module separate file with reusable code
# see documentation
# import math
# print(math.floor(2.9))

# if statements ------------------------------

# is_hot = False
# is_cold = True

# if is_hot:
# print("It's a hot day")
# print("Drink plenty of water")
# elif is_cold:
# print("It's a cold day")
# print("Wear warm clothes")
# else:
# print("It's a lovely day")
# print("Enjoy your day")

# exercise
# price = 1_000_000
# has_good_credit = True

# if has_good_credit:
# down_payment = .1 * price
# else:
# down_payment = .1 * price
# print(f"Down payment: ${down_payment}")

name = "Clive Madungwe"

# if len(name) < 3:
# print("Name must be at least 3 characters: ")
# elif len(name) > 50:
# print("Name must be a maximum of 50 characters: ")
# else:3
# print("Names looks good")

# while loops-------------------------------------------------------

# i = 1
# while i <= 5:
# print('*' * i)
# i = i + 1
# print("Done")

# Guess game

# secret_number = 9
# guess_count = 0
# guess_limit = 3
# num_of_guess_left = 0
# while guess_count < guess_limit:
# guess = int(input("Guess: "))
# guess_count += 1
# num_of_guess_left = 3 - guess_limit
# if guess == secret_number:
# print("You won: ")
# break
# else:
# print(f" You have {num_of_guess_left} guess left")
# print("Sorry you failed")

# car game--------

# command = ""
# started = False
# while True:
#     command = input("> ").lower()
#     if command == "start":
#         if started:
#             print("Car is already started: ")
#         else:
#             started = True
#             print("Car started...")
#     elif command == "stop":
#         if not started:
#             print("Car is already stopped!" )
#         else:
#             started = False
#             print("Car stopped.")
#     elif command == "help":
#         print("""
#         start - to start the car
#         stop - to stop the car
#         quit - to quit
#         """)
#     elif command == "quit":
#         break
#     else:
#          print("Sorry, I don't understand that! ")

# for item in 'Python':
#     print(item)

# for item in range(1, 10):
#     print(item)

# for x in range(4):
#     for y in range(3):
#         print(f"({x}, {y}) ")

# numbers =[5, 2, 5, 2, 2]
# for x_count in numbers:
#     print("x" * x_count)

# lists---------------------------------------------------------

# names = ["John", "Bob", "Mosh", "Sarah", "Mary"]
# names[0] = "Jon"
# print(names[0])

# numbers = [3, 6, 2, 8, 4, 10]
# max = numbers[0]
# for number in numbers:
#     if number > max:
#         max = number
# print(max)

# 2D lists----------------------------------------------------
# matrix = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [7, 8, 9]
# ]
# print(matrix[0][1])

# matrix = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [7, 8, 9]
# ]
# for row in matrix:
#     for item in row:
#         print(matrix)

# list methods copy,append, insert,clear,pop,index,in(print(5 in numbers)),sort,reverse,copy e.g numbers.append(5)
# numbers = [2, 2, 4, 6, 3, 4, 6, 1]
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# turples are immutable-----------------------------(numbers = (1, 2, 3)

# coordinates = (1, 2, 3)
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]
#
# unpaking
#
# x, y, z = coordinates

# Dictionaries----------------------------------------------
# customer = {
#     "name": "Clive Madungwe",\
#     "age": 30,
#     "is_verified": True
# }
# print(customer['name'])
# print(customer.get("birthdate", "April 5 1997"))

# phone = input("Phone: ")
# digits_mapping = {
#     "1": "One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four"
# }
# output = ""
# for ch in phone:
#     output += digits_mapping.get(ch, "!") + " "
# print(output)

# Emoji Converter
# message = input(">")
# words = message.split(" ")
# emojis = {
#     ":)": "😀",
#     ":(": "😔"
# }
# output = ""
# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)

# functions------------------------------------------------------

# def greet_user(name):
#     print(f"Hi {name}!")
#     print("Welcome aboard")
#
#
# print("Start")
# greet_user("Clive")
# print("Finish")

# return statement
# def square(number):
#     return number * number
#
#
# result = square(3)
# print(result)
# print(square(3))

# reusable functions
# def emoji_converter(message):
#     words = message.split(" ")
#     emojis = {
#         ":)": "😀",
#         ":(": "😔"
#     }
#     output = ""
#     for word in words:
#         output += emojis.get(word, word) + " "
#     return output
#
#
# message = input(">")
# print(emoji_converter(message))

# Exceptions---------------------------------------------
# try:
#     age = int(input("Age: "))
#     income = 2000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print("Age cannot be 0")
# except ValueError:
#     print("Invalid value")

# Classes
# class Point:
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")
#
#
# point1 = Point()
# point1.x = 10
# point1.y = 20
# print(point1.x)
# point1.draw()
#
# point2 = Point()

# constructor
# class Point:
#     def __int__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")
#
#
# point = Point(10, 20)
# print(point.x)

# class Person:
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f"hi, I am {self.name}")
#
#
# clive = Person("Clive Madungwe")
# clive.talk()
#
# trisher = Person(" Trisher Mpatsi")
# trisher.talk()

# inheritance--------------------------------------------------------
# class Mammal:
#     def walk(self):
#         print("walk")
#
#
# class Dog(Mammal):
#     def bark(self):
#         print("bark")
#
#
# class Cat(Mammal):
#     def be_annoying(self):
#         print("annoying")
#
#
# dog1 = Dog()
# dog1.walk()

# modules -------------------------------
# open new file converters.py
# import converters
# print(converters.kg_to_lbs(70))
# from converters import kg_to lbs

# from ecommerce.shipping import calc_shipping
#
# calc_shipping()

# modules random

# import random
#
# for i in range(3):
#     print(random.random())
#     print(random.randint(10, 20))

# import random
#
# members = ['Clive', 'Bitton', '25', 'DB']
# leader = random.choice(members)
# print(leader)


# import random
#
#
# class Dice:
#     def roll(self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#         return first, second
#
#
# dice = Dice()
# print(dice.roll())
#
# from pathlib import Path
#
# path = Path("ecommerce")
# print(path.exists())
# print(path.mkdir())
# print(path.glob("*.py"))

# PyPI

# automation
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]
cell = sheet.cell(1, 1)
# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * .9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save("transaction2.xlsx")

